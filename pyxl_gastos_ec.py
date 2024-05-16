from openpyxl import workbook, load_workbook
from tkinter import StringVar, Tk, ttk, Label, Button

wb = load_workbook('C:/Users/ederc/OneDrive/Documentos/Finance/finanzas_familiares/IngresosyGastos_2024/IngresosyGastos_2024.xlsx')
ws = wb.active

wg = load_workbook('C:/Users/ederc/OneDrive/Documentos/Finance/finanzas_familiares/IngresosyGastos_2024/Gastos_excel/Abril/excel_abril_2024.xlsx')
wt = wg.active

dates = []
store = []
quantity = []

for col_cells in wt.iter_cols(min_col=1, max_col=1):
    cl = 0
    cval = ''
    for cell in col_cells:
        if cl < 1:
            None
        else:
            if cell.value == None:
                break
            else:
                cval = cell.value
                dates.append(cval)                
        cl += 1
#print(dates)

for col_cells in wt.iter_cols(min_col=2, max_col=2):
    c_spent = 0
    cval_spent = ''
    for cell in col_cells:
        if c_spent < 1:
            None
        else:
            if cell.value == None:
                break
            else:
                cval_spent = cell.value
                store.append(cval_spent)
        c_spent += 1
#print(spents)

for col_cells in wt.iter_cols(min_col=4, max_col=4):
    c_purch = 0
    cval_purch = ''
    for cell in col_cells:
        #print(cell.value)
        if c_purch < 1:
            None
        else:
            if cell.value == None:
                #print(cell.value)
                break
            else:
                #print(cell.value)
                cval_purch = cell.value
                #cval_purch = float(cval_purch)
                quantity.append(cval_purch)
        c_purch += 1

cd = 3
for col_cells in dates:
    if cell.value == None:
        break
    else:
        ws['E'+ str(cd)] = col_cells               
    cd += 1

cntr_store = 3
for col_cells in store:
    if cell.value == None:
        break
    else:
        ws['F'+str(cntr_store)] = col_cells
    cntr_store += 1

cntr_quantity = 3
for col_cells in quantity:
    if cell.value == None:
        break
    else:
        ws['I'+str(cntr_quantity)] = col_cells
    cntr_quantity += 1



myDict = {'Servicios': ['NATURGY', 'AGUAYDRENAJE', 'MESES', 'CFE', 'AGUA DREN'], 
          'Servicios_Entretenimiento': ['IZZI', 'SKY', 'NETFLIX', 'APPLE', 'SPOTIFY', 'PRIME'],
           'Mandado': ['CORNERSHOP', 'HEB', 'SAMS', 'SORIANA', 'JUSTOPJU', 'JUSTO'], 'Farmacia': ['FARM GUADALAJARA', 'FARM GUAD' 'BENAVIDES'], 
           'Fijo': ['MERPAGO', 'INTERES', 'Citibanamex', 'A MESES', 'PROTEGIDO', 'MERCADO PAGO'],
            'Conveniencia': ['7 ELEVEN', '7ELEVEN', 'OXXO', 'NAYAX'], 'Salud': ['CMSH', 'MEDICENTRO', 'GINECO', 'DR QUEEN'], 
            'Comidas': ['UBER EAT', 'RAPPI', 'CARLS JR'],
            'Transportacion': ['GAS', 'PETRO', 'TRIPUPM', 'UBER TRI'], 'Anual': ['TRADINGVIEW', 'DISNEY', 'NINTENDO', 'MUNICIPIO STA CATARINA'], 
            'Pedidos': ['MARKETPLACE', 'MX ANE', 'AMAZON MX MARKETPLACE'], 'Imprevistos': ['PASTELERIA', 'HOMEDEPOT'], 'Diversion':['Estadio', 'PLAYTICA', 'diversion_cash', 'PARKIT', 'HELADOS SULTANA'],
            'Abarrotes': ['abarrotes_cash', 'cerveza_cash'], 'Ropa': ['ropa_cash'], 'Mejoras_casa':['mejoras_casa_cash'], 'Educacion': ['educacion_cash']}

key_words  = []
words_list = myDict.values()
for z in words_list:
    for y in z:
        key_words.append(y)


def seek_dict_key(KY, dict):
    '''KY Will be the key word to look,
    and dict will be a dictionary with all the keys and values'''
    myKey = None
    for i in dict:
        if KY in dict[i]:
            #print(i)
            myKey = i
        else:
            myKey = myKey
    return myKey


def find_word(cell_data, KW):
    ''' cell data is the data contained in the current cell, and KW is the key words'''
    occur = 0
    for j in KW:
        tamano = len(cell_data)
        looklen = len(j)
        #print(j)
        for i in range(tamano):
            if occur == 1:
                break
            else:
                reb = cell_data[i:(looklen+i)]
                #print(reb)
                #print (j)
                if reb == j:
                    occur += 1
                else:
                    occur = occur
    return (str(reb))

def ask_expense_type(cell_value):
    def set_value():
        selected_value.set(expense_type_combobox.get())
        dialog.destroy()

    dialog = Tk()
    dialog.title("Seleccionar tipo de gasto")

    selected_value = StringVar()
    label = Label(dialog, text=f"¿Qué tipo de gasto es este? {cell_value}")
    label.pack()

    expense_type_combobox = ttk.Combobox(dialog, values=list(myDict.keys()))
    expense_type_combobox.pack()
    expense_type_combobox.current(0)

    button = Button(dialog, text="Ok", command=set_value)
    button.pack()

    dialog.mainloop()
    return selected_value.get()


for col_cells in ws.iter_cols(min_col=6, max_col=6):
    cl = 1
    categoria = ''
    for cell in col_cells:
        if cl <= 2:
            None
        else:
            if cell.value == None:
                break
            else:
                categoria = seek_dict_key(find_word(cell.value, key_words), myDict)
                if categoria == None:
                    #L1 = Label(text="Que tipo de gasto es este: " + cell.value)
                    #categoria = input("Que tipo de gasto es este" + cell.value + ':')
                    categoria = ask_expense_type(cell.value)
                    ws['H'+ str(cl)] = categoria
                else:
                    ws['H'+ str(cl)] = categoria
                    #print(categoria)                
        cl += 1


Dict_values = {'Servicios': [], 'Servicios_Entretenimiento': [],'Mandado': [], 'Farmacia': [], 
           'Fijo': [],'Conveniencia': [], 'Salud': [], 'Comidas': [],
            'Transportacion': [], 'Anual': [], 'Pedidos': [], 'Imprevistos': [], 'Diversion': [], 
            'Abarrotes': [], 'Ropa': [], 'Mejoras_casa': [], 'Educacion': []}

for col_cells in ws.iter_cols(min_col=8, max_col=8):
    cntr_H = 1
    for cell in col_cells:
        if cntr_H <= 2:
            None
        else:
            if cell.value == None:
                break
            else:
                if cell.value == 'Servicios':
                    Dict_values['Servicios'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Servicios_Entretenimiento':
                    Dict_values['Servicios_Entretenimiento'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Mandado':
                    Dict_values['Mandado'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Farmacia':
                    Dict_values['Farmacia'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Fijo':
                    Dict_values['Fijo'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Conveniencia':
                    Dict_values['Conveniencia'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Salud':
                    Dict_values['Salud'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Comidas':
                    Dict_values['Comidas'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Transportacion':
                    Dict_values['Transportacion'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Anual':
                    Dict_values['Anual'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Pedidos':
                    Dict_values['Pedidos'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Imprevistos':
                    Dict_values['Imprevistos'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Diversion':
                    Dict_values['Diversion'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == 'Abarrotes':
                    Dict_values['Abarrotes'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == ['Ropa']:
                    Dict_values['Ropa'].append(ws['I'+ str(cntr_H)].value)
                elif cell.value == ['Mejoras_casa']:
                    Dict_values['Mejoras_casa'].append(ws['I'+ str(cntr_H)].value)
                else:
                    Dict_values['Educacion'].append(ws['I'+ str(cntr_H)].value) 
        cntr_H += 1
suma = 0
for k in Dict_values['Servicios']:
    suma = suma + k
ws['L14'] = suma

suma = 0
for k in Dict_values['Servicios_Entretenimiento']:
    suma = suma + k
ws['L15'] = suma

suma = 0
for k in Dict_values['Mandado']:
    suma = suma + k
ws['L11'] = suma

suma = 0
for k in Dict_values['Farmacia']:
    suma = suma + k
ws['L5'] = suma

suma = 0
for k in Dict_values['Fijo']:
    suma = suma + k
ws['L4'] = suma

suma = 0
for k in Dict_values['Conveniencia']:
    suma = suma + k
ws['L3'] = suma

suma = 0
for k in Dict_values['Salud']:
    suma = suma + k
ws['L6'] = suma

suma = 0
for k in Dict_values['Comidas']:
    suma = suma + k
ws['L8'] = suma

suma = 0
for k in Dict_values['Transportacion']:
    suma = suma + k
ws['L7'] = suma

suma = 0
for k in Dict_values['Anual']:
    suma = suma + k
ws['L13'] = suma

suma = 0
for k in Dict_values['Pedidos']:
    suma = suma + k
ws['L9'] = suma

suma = 0
for k in Dict_values['Imprevistos']:
    suma = suma + k
ws['L10'] = suma

suma = 0
for k in Dict_values['Diversion']:
    suma = suma + k
ws['L12'] = suma

suma = 0
for k in Dict_values['Abarrotes']:
    suma = suma + k
ws['L18'] = suma

suma = 0
for k in Dict_values['Ropa']:
    suma = suma + k
ws['L17'] = suma

suma = 0
for k in Dict_values['Mejoras_casa']:
    suma = suma + k
ws['L16'] = suma

suma = 0
for k in Dict_values['Educacion']:
    suma = suma + k
ws['L19'] = suma

print(Dict_values)
wb.save('C:/Users/ederc/OneDrive/Documentos/Finance/finanzas_familiares/IngresosyGastos_2024/IngresosyGastos_2024.xlsx')